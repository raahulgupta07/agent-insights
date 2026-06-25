import csv
import io
import re
import uuid
from datetime import datetime, timezone, timedelta
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.models.organization import Organization
from app.models.membership import Membership
from app.schemas.organization_schema import OrganizationCreate, OrganizationSchema, OrganizationAndRoleSchema, OrganizationUpdate
from app.schemas.organization_schema import MembershipCreate, MembershipSchema, MembershipUpdate
from app.schemas.organization_schema import (
    MembershipImportRow,
    MembershipImportSummary,
    MembershipImportReport,
    MEMBERSHIP_NOTE_MAX_LENGTH,
)
from app.schemas.organization_settings_schema import OrganizationSettingsCreate
from app.services.organization_settings_service import OrganizationSettingsService
from app.schemas.user_schema import UserSchema
from uuid import UUID
from app.models.user import User
from typing import List
from sqlalchemy.orm import selectinload
from fastapi import HTTPException
from sqlalchemy import delete
from app.services.llm_service import LLMService
from app.services.test_suite_service import TestSuiteService
from app.settings.config import settings
from fastapi import Request
from typing import Optional
from app.settings.logging_config import get_logger
from app.core.telemetry import telemetry

logger = get_logger(__name__)

# How long an invite link stays valid. Resending rotates the token and resets
# this window.
INVITE_EXPIRY_DAYS = 14


class OrganizationService:

    def __init__(self):
        self.llm_service = LLMService()
        self.organization_settings_service = OrganizationSettingsService()
        self.test_suite_service = TestSuiteService()
    async def create_organization(self, db: AsyncSession, organization_data: OrganizationCreate, current_user: User) -> OrganizationSchema:

        total_orgs = await db.execute(select(Organization))
        total_orgs = total_orgs.scalars().all().__len__()
        if total_orgs > 0 and not settings.dash_config.features.allow_multiple_organizations:
            raise HTTPException(status_code=400, detail="You cannot create more than one organization")
        
        organization = Organization(**organization_data.dict())
        db.add(organization)
        await db.commit()
        await db.refresh(organization)

        # Telemetry: organization created
        try:
            await telemetry.capture(
                "organization_created",
                {
                    "organization_id": str(organization.id),
                    "name_length": len((organization.name or "").strip()),
                },
                user_id=current_user.id,
                org_id=organization.id,
            )
        except Exception:
            pass

        await self.organization_settings_service.create_default_settings(db, organization, current_user)
        await self.add_member(db, MembershipCreate(role="admin", user_id=current_user.id, organization_id=organization.id), current_user)
        await self.llm_service.set_default_models_from_config(db, organization, current_user)
        await self.test_suite_service.ensure_default_for_org(db, organization.id, current_user)

        # Create RBAC role_assignment for the admin system role
        await self._assign_system_role(db, organization.id, str(current_user.id), "admin")

        return OrganizationSchema.from_orm(organization)

    async def _assign_system_role(self, db: AsyncSession, org_id: str, user_id: str, role_name: str) -> None:
        """Assign a system role to a user via role_assignments (RBAC path)."""
        from app.models.role import Role
        from app.models.role_assignment import RoleAssignment

        try:
            result = await db.execute(
                select(Role).where(
                    Role.name == role_name,
                    Role.is_system == True,
                    Role.organization_id.is_(None),
                    Role.deleted_at.is_(None),
                )
            )
            system_role = result.scalar_one_or_none()
            if not system_role:
                return
            # Check if assignment already exists
            existing = await db.execute(
                select(RoleAssignment).where(
                    RoleAssignment.organization_id == org_id,
                    RoleAssignment.role_id == system_role.id,
                    RoleAssignment.principal_type == "user",
                    RoleAssignment.principal_id == user_id,
                )
            )
            if existing.scalar_one_or_none():
                return
            assignment = RoleAssignment(
                organization_id=org_id,
                role_id=system_role.id,
                principal_type="user",
                principal_id=user_id,
            )
            db.add(assignment)
            await db.commit()
        except Exception:
            await db.rollback()
            # Don't break org creation if RBAC tables don't exist yet (pre-migration)

    async def get_organization(self, db: AsyncSession, organization_id: str, current_user: User) -> OrganizationSchema:
        result = await db.execute(select(Organization).where(Organization.id == organization_id))
        return result.scalar_one_or_none()
    
    async def get_members(self, db: AsyncSession, organization: Organization, current_user: User) -> List[MembershipSchema]:
        from app.models.role_assignment import RoleAssignment
        from app.models.role import Role
        from app.models.group_membership import GroupMembership
        from app.models.group import Group
        from app.schemas.organization_schema import RoleSummarySchema

        result = await db.execute(
            select(Membership)
            .options(selectinload(Membership.user))
            .where(Membership.organization_id == organization.id)
        )
        memberships = result.scalars().all()

        schemas = []
        for membership in memberships:
            schema = MembershipSchema.from_orm(membership)
            if membership.user_id:
                # Registered user: direct ('user' principal) + group-inherited.
                schema.roles = await self._resolve_member_roles(
                    db, organization.id,
                    direct_principal_type="user",
                    direct_principal_id=membership.user_id,
                    group_filter=GroupMembership.user_id == membership.user_id,
                )
            else:
                # Pending invite: direct ('membership' principal) +
                # group-inherited via pending group memberships.
                schema.roles = await self._resolve_member_roles(
                    db, organization.id,
                    direct_principal_type="membership",
                    direct_principal_id=membership.id,
                    group_filter=GroupMembership.membership_id == membership.id,
                )
            schemas.append(schema)
        return schemas

    async def _resolve_member_roles(
        self, db: AsyncSession, organization_id: str,
        direct_principal_type: str, direct_principal_id: str, group_filter,
    ) -> List["RoleSummarySchema"]:
        """Resolve a member's direct + group-inherited roles.

        Works for both registered users (principal_type='user', groups joined
        by user_id) and pending invites (principal_type='membership', groups
        joined by membership_id).
        """
        from app.models.role_assignment import RoleAssignment
        from app.models.group_membership import GroupMembership
        from app.models.group import Group
        from app.schemas.organization_schema import RoleSummarySchema

        # Direct role assignments
        ra_result = await db.execute(
            select(RoleAssignment)
            .options(selectinload(RoleAssignment.role))
            .where(
                RoleAssignment.organization_id == organization_id,
                RoleAssignment.principal_type == direct_principal_type,
                RoleAssignment.principal_id == direct_principal_id,
                RoleAssignment.deleted_at.is_(None),
            )
        )
        assignments = ra_result.scalars().all()
        roles = [
            RoleSummarySchema(id=a.role.id, name=a.role.name, source="direct")
            for a in assignments if a.role
        ]

        # Group-inherited role assignments
        gm_result = await db.execute(
            select(GroupMembership.group_id, Group.name)
            .join(Group, Group.id == GroupMembership.group_id)
            .where(
                group_filter,
                Group.organization_id == organization_id,
                GroupMembership.deleted_at.is_(None),
                Group.deleted_at.is_(None),
            )
        )
        member_groups = gm_result.all()  # [(group_id, group_name), ...]

        if member_groups:
            group_ids = [g[0] for g in member_groups]
            group_names = {g[0]: g[1] for g in member_groups}
            seen_role_ids = {r.id for r in roles}

            group_ra_result = await db.execute(
                select(RoleAssignment)
                .options(selectinload(RoleAssignment.role))
                .where(
                    RoleAssignment.organization_id == organization_id,
                    RoleAssignment.principal_type == "group",
                    RoleAssignment.principal_id.in_(group_ids),
                    RoleAssignment.deleted_at.is_(None),
                )
            )
            for a in group_ra_result.scalars().all():
                if a.role and a.role.id not in seen_role_ids:
                    group_name = group_names.get(a.principal_id, "unknown")
                    roles.append(RoleSummarySchema(
                        id=a.role.id,
                        name=a.role.name,
                        source=f"group:{group_name}",
                    ))
                    seen_role_ids.add(a.role.id)

        return roles
    
    async def get_member(self, db: AsyncSession, membership_id: str, organization_id: str, current_user: User) -> MembershipSchema:
        result = await db.execute(
            select(Membership)
            .options(selectinload(Membership.user))
            .where(Membership.id == membership_id, Membership.organization_id == organization_id)
        )
        return result.scalar_one_or_none()
    

    async def _count_org_memberships(self, db: AsyncSession, organization_id) -> int:
        """Count all memberships in an org — active members and pending invites alike.

        Pending invites (user_id is NULL) count too, so the license seat cap can't be
        bypassed by leaving invites unaccepted.
        """
        from sqlalchemy import func
        result = await db.execute(
            select(func.count(Membership.id)).where(
                Membership.organization_id == organization_id
            )
        )
        return result.scalar() or 0

    async def _enforce_user_limit(self, db: AsyncSession, organization_id, adding: int = 1) -> None:
        """Raise 402 if adding `adding` member(s) would exceed the license seat cap.

        No-op when unlicensed/unset (max_users == -1 → unlimited).
        """
        from app.ee.license import get_max_users
        max_users = get_max_users()
        if max_users < 0:
            return
        current = await self._count_org_memberships(db, organization_id)
        if current + adding > max_users:
            raise HTTPException(
                status_code=402,
                detail=(
                    f"User limit reached for your license ({max_users}). "
                    "Contact sales to increase your seat count."
                ),
            )

    async def add_member(self, db: AsyncSession, membership_data: MembershipCreate, current_user: User) -> MembershipSchema:
        #check if email is already a user
        # if it is, add user_id to membership and remove email
        # then, check if user (or email) already maps to a membership in this organization
        # if it does, raise an error
        membership_exists = await self._is_email_already_in_organization(db, membership_data.email, membership_data.organization_id)
        if membership_exists:
            raise HTTPException(status_code=400, detail="Already a member with this email")

        # Enforce per-organization seat cap from the enterprise license (if any).
        # Checked after the duplicate guard so re-adding an existing email still
        # 400s rather than being masked by a seat-limit 402.
        await self._enforce_user_limit(db, membership_data.organization_id)
        
        user = await db.execute(select(User).where(User.email == membership_data.email))
        user = user.scalar_one_or_none()

        # Store the email for invitation before potentially setting it to None
        invitation_email = membership_data.email

        if user:
            membership_data.user_id = user.id
            membership_data.email = None

        membership = Membership(**membership_data.dict())
        # Pending (unregistered) invite → stamp a 14-day expiry on the invite
        # link. invite_token is auto-generated by the model default.
        if membership.user_id is None:
            membership.invite_expires_at = datetime.utcnow() + timedelta(days=INVITE_EXPIRY_DAYS)

        db.add(membership)
        await db.commit()
        await db.refresh(membership)
        
        # Reload the membership with the user relationship
        result = await db.execute(
            select(Membership)
            .options(selectinload(Membership.user))
            .where(Membership.id == membership.id)
        )
        membership_with_user = result.scalar_one()
        # Telemetry: organization member invited/added
        try:
            await telemetry.capture(
                "organization_member_added",
                {
                    "organization_id": str(membership_with_user.organization_id),
                    "membership_id": str(membership_with_user.id),
                    "role": membership_with_user.role,
                    "user_id": str(membership_with_user.user_id) if membership_with_user.user_id else None,
                },
                user_id=current_user.id,
                org_id=membership_with_user.organization_id,
            )
        except Exception:
            pass
        
        # Send invitation email immediately, but reliably: awaited (so we know
        # the real outcome), retried on transient SMTP errors, and timeout-bounded
        # so a hung relay can't block the request. The outcome is surfaced on the
        # response so the admin UI can warn instead of silently "succeeding".
        # Only pending invites get a sign-up email (existing users already have
        # an account). The link carries the invite token so the recipient can
        # be verified at registration time.
        invite_email_status: Optional[str] = None
        if invitation_email and membership_with_user.user_id is None:
            if not (hasattr(settings, 'email_client') and settings.email_client):
                invite_email_status = "skipped_no_smtp"
            else:
                invite_email_status = await self._send_invitation_email(
                    invitation_email, membership_with_user.invite_token
                )

        # Create RBAC role_assignment if user_id is set
        if membership_with_user.user_id and membership_data.role:
            await self._assign_system_role(db, membership_data.organization_id, membership_with_user.user_id, membership_data.role)

        schema = MembershipSchema.from_orm(membership_with_user)
        schema.invite_email_status = invite_email_status
        return schema

    async def create_user_with_password(
        self,
        db: AsyncSession,
        organization_id: str,
        email: str,
        password: str,
        name: str,
        role: str,
        current_user: User,
    ) -> MembershipSchema:
        """Admin-side DIRECT user creation — no email invite.

        Creates a real, active, verified User with the given password and adds
        it to this organization with `role`. The user can log in immediately.
        If the email already maps to a user, that user is reused (and added to
        the org if not already a member). Idempotent-ish: a duplicate member
        400s rather than silently re-adding.
        """
        email = (email or "").strip()
        if not email:
            raise HTTPException(status_code=400, detail="Email is required")

        # Already a member of THIS org (pending invite or active) → 400.
        if await self._is_email_already_in_organization(db, email, organization_id):
            raise HTTPException(status_code=400, detail="Already a member with this email")

        # Seat cap (enterprise license), same guard as add_member.
        await self._enforce_user_limit(db, organization_id)

        # Does a global user already exist for this email?
        existing = (
            await db.execute(select(User).where(User.email == email))
        ).scalar_one_or_none()

        if existing is not None:
            user_id = str(existing.id)
        else:
            # Create the user THROUGH the fastapi-users manager so the standard
            # on_after_register hook fires. This is NOT the first user (the org
            # already exists), so the hook does NOT auto-create an org — we
            # attach the membership to THIS org explicitly below.
            from app.dependencies import async_session_maker, get_user_db
            from app.core.auth import get_user_manager
            from app.schemas.user_schema import UserCreate

            async def _aclose(gen):
                try:
                    await gen.__anext__()
                except StopAsyncIteration:
                    pass
                except Exception:
                    pass

            async with async_session_maker() as mgr_session:
                user_db_gen = get_user_db(session=mgr_session)
                user_db = await user_db_gen.__anext__()
                try:
                    manager_gen = get_user_manager(user_db=user_db)
                    manager = await manager_gen.__anext__()
                    try:
                        try:
                            created = await manager.create(
                                UserCreate(email=email, password=password, name=name),
                                safe=False,
                            )
                        except Exception as exc:
                            raise HTTPException(
                                status_code=400,
                                detail=f"Could not create user: {exc}",
                            )
                        user_id = str(created.id)
                    finally:
                        await _aclose(manager_gen)
                finally:
                    await _aclose(user_db_gen)

            # Make the account usable right away (no email-verification step).
            elevate = (
                await db.execute(select(User).where(User.id == user_id))
            ).scalar_one_or_none()
            if elevate is not None:
                elevate.is_active = True
                elevate.is_verified = True
                await db.commit()

        # Attach the membership to this org with a resolved user_id (no invite).
        membership = Membership(organization_id=organization_id, user_id=user_id, role=role)
        db.add(membership)
        await db.commit()
        await db.refresh(membership)

        # RBAC role assignment (mirrors add_member).
        if role:
            await self._assign_system_role(db, organization_id, user_id, role)

        result = await db.execute(
            select(Membership)
            .options(selectinload(Membership.user))
            .where(Membership.id == membership.id)
        )
        membership_with_user = result.scalar_one()

        try:
            await telemetry.capture(
                "organization_member_created_direct",
                {
                    "organization_id": str(organization_id),
                    "membership_id": str(membership_with_user.id),
                    "role": role,
                    "user_id": user_id,
                },
                user_id=current_user.id,
                org_id=organization_id,
            )
        except Exception:
            pass

        return MembershipSchema.from_orm(membership_with_user)

    async def get_user_organizations(self, db: AsyncSession, current_user: User) -> List[OrganizationAndRoleSchema]:
        from app.core.permission_resolver import resolve_permissions

        result = await db.execute(
            select(Organization, Membership.role)
            .join(Membership)
            .where(Membership.user_id == current_user.id)
        )
        results = result.all()
        org_ids = [org.id for org, _ in results]
        # Load settings for these orgs to extract icon_url
        from app.models.organization_settings import OrganizationSettings
        settings_map = {}
        if org_ids:
            sres = await db.execute(select(OrganizationSettings).where(OrganizationSettings.organization_id.in_(org_ids)))
            for s in sres.scalars().all():
                settings_map[s.organization_id] = s

        formatted = []
        from app.services.usage_policy_service import usage_policy_service
        for org, role in results:
            icon_url = None
            ai_analyst_name = "City Agent Insights"  # Default value
            org_settings = settings_map.get(org.id)
            if org_settings and isinstance(org_settings.config, dict):
                general = org_settings.config.get('general') or {}
                icon_url = general.get('icon_url')
                ai_analyst_name = general.get('ai_analyst_name') or "City Agent Insights"

            # Resolve RBAC permissions
            resolved = await resolve_permissions(db, str(current_user.id), str(org.id))

            # Build resource_permissions dict for frontend
            resource_perms = {}
            for (res_type, res_id), perms in resolved.resource_permissions.items():
                key = f"{res_type}:{res_id}"
                resource_perms[key] = sorted(perms)

            # Check enterprise license
            from app.ee.license import has_feature
            is_enterprise = has_feature("custom_roles")
            usage_quota = await usage_policy_service.get_user_quota_summary(
                db,
                str(org.id),
                str(current_user.id),
            )

            formatted.append(OrganizationAndRoleSchema(
                id=org.id,
                name=org.name,
                description=org.description,
                role=role,  # backward compat
                roles=resolved.role_names,
                permissions=sorted(resolved.org_permissions),
                resource_permissions=resource_perms,
                is_enterprise=is_enterprise,
                icon_url=icon_url,
                ai_analyst_name=ai_analyst_name,
                usage_quota=usage_quota,
            ))
        return formatted


    async def remove_member(self, db: AsyncSession, organization_id, membership_id: str, current_user: User, organization: Organization) -> None:
        from app.core.permission_resolver import assert_full_admin_exists

        membership = await self.get_member(db, membership_id, organization_id, current_user)
        if not membership:
            raise HTTPException(status_code=404, detail="Membership not found")
        if membership.user_id:
            # RBAC lockout prevention: ensure at least one user keeps full_admin_access
            await assert_full_admin_exists(db, organization_id, exclude_user_id=membership.user_id)
        else:
            # Pending invite: clean up any pre-assigned RBAC keyed by this
            # membership. Done explicitly (not via FK cascade) so it holds on
            # SQLite, where foreign-key enforcement is off by default.
            from app.models.role_assignment import RoleAssignment
            from app.models.group_membership import GroupMembership
            from app.models.usage_policy import UsagePolicyAssignment
            await db.execute(
                delete(RoleAssignment).where(
                    RoleAssignment.organization_id == organization_id,
                    RoleAssignment.principal_type == "membership",
                    RoleAssignment.principal_id == membership_id,
                )
            )
            await db.execute(
                delete(GroupMembership).where(GroupMembership.membership_id == membership_id)
            )
            await db.execute(
                delete(UsagePolicyAssignment).where(
                    UsagePolicyAssignment.organization_id == organization_id,
                    UsagePolicyAssignment.principal_type == "membership",
                    UsagePolicyAssignment.principal_id == membership_id,
                )
            )

        await db.execute(delete(Membership).where(Membership.id == membership_id))
        await db.commit()
    
    async def update_member(self, db: AsyncSession, membership_id: str, organization_id: str, membership_data: MembershipUpdate, current_user: User, organization: Organization) -> MembershipSchema:
        from app.core.permission_resolver import assert_full_admin_exists

        membership = await self.get_member(db, membership_id, organization_id, current_user)
        if not membership:
            raise HTTPException(status_code=404, detail="Membership not found")

        # RBAC lockout prevention: whenever we mutate a membership's role field,
        # verify at least one user still holds full_admin_access in the org.
        if membership.user_id and membership_data.role and membership_data.role != membership.role:
            await assert_full_admin_exists(db, organization_id, exclude_user_id=membership.user_id)

        update = membership_data.dict(exclude_unset=True)
        if "role" in update and update["role"] is not None:
            membership.role = update["role"]
        if "note" in update:
            membership.note = update["note"]
        await db.commit()

        # Reload with the user relationship so MembershipSchema serialization
        # doesn't trigger an async lazy-load in a session that's already done.
        result = await db.execute(
            select(Membership)
            .options(selectinload(Membership.user))
            .where(Membership.id == membership.id)
        )
        return MembershipSchema.from_orm(result.scalar_one())

    async def resend_invite(self, db: AsyncSession, membership_id: str, organization_id: str) -> MembershipSchema:
        """Rotate the invite token, reset the 14-day expiry, and re-send the email.

        Pending invites only (a registered member has no invite to resend). The
        old link stops working as soon as the token is rotated.
        """
        membership = await self.get_member(db, membership_id, organization_id, None)
        if not membership:
            raise HTTPException(status_code=404, detail="Membership not found")
        if membership.user_id is not None or not membership.email:
            raise HTTPException(status_code=400, detail="This member has already registered; nothing to resend")

        membership.invite_token = str(uuid.uuid4())
        membership.invite_expires_at = datetime.utcnow() + timedelta(days=INVITE_EXPIRY_DAYS)
        await db.commit()
        await db.refresh(membership)

        status = None
        if hasattr(settings, 'email_client') and settings.email_client:
            status = await self._send_invitation_email(membership.email, membership.invite_token)
        else:
            status = "skipped_no_smtp"

        result = await db.execute(
            select(Membership).options(selectinload(Membership.user)).where(Membership.id == membership.id)
        )
        schema = MembershipSchema.from_orm(result.scalar_one())
        schema.invite_email_status = status
        return schema

    async def get_invite_link(self, db: AsyncSession, membership_id: str, organization_id: str) -> dict:
        """Return the tokenized sign-up link for a pending invite (admin use).

        Lets an admin copy/share the link directly (handy when SMTP is off) and
        is the proof-of-invite the recipient presents at registration. If the
        invite has already expired, the token is regenerated and the 14-day
        window reset so the copied link is always usable (no email is sent —
        that's what Resend does). A still-valid link is returned untouched so we
        don't invalidate one that was already emailed.
        """
        from urllib.parse import quote

        membership = await self.get_member(db, membership_id, organization_id, None)
        if not membership:
            raise HTTPException(status_code=404, detail="Membership not found")
        if membership.user_id is not None or not membership.email:
            raise HTTPException(status_code=400, detail="This member has already registered; no invite link")

        expires = membership.invite_expires_at
        regenerated = False
        if not membership.invite_token or (expires is not None and expires < datetime.utcnow()):
            membership.invite_token = str(uuid.uuid4())
            membership.invite_expires_at = datetime.utcnow() + timedelta(days=INVITE_EXPIRY_DAYS)
            await db.commit()
            await db.refresh(membership)
            regenerated = True

        token = membership.invite_token
        url = (
            f"{settings.dash_config.base_url}/users/sign-up"
            f"?token={quote(token or '')}&email={quote(membership.email)}"
        )
        return {
            "token": token,
            "email": membership.email,
            "url": url,
            "invite_expires_at": membership.invite_expires_at,
            "regenerated": regenerated,
        }

    async def update_organization(self, db: AsyncSession, organization: Organization, data: OrganizationUpdate, current_user: User) -> OrganizationSchema:
        """Update organization basic fields like name/description."""
        update = data.dict(exclude_unset=True)
        if 'name' in update and update['name']:
            organization.name = update['name']
        if 'description' in update:
            organization.description = update['description']
        await db.commit()
        await db.refresh(organization)
        return OrganizationSchema.from_orm(organization)
    
    async def _is_email_already_in_organization(self, db: AsyncSession, email: str, organization_id: str) -> bool:
        user = await db.execute(select(User).where(User.email == email))
        user = user.scalar_one_or_none()
        if user:
            membership = await db.execute(select(Membership).where(Membership.user_id == user.id, Membership.organization_id == organization_id))
            membership = membership.scalar_one_or_none()
            return membership 
        
        email_membership = await db.execute(select(Membership).where(Membership.email == email, Membership.organization_id == organization_id))
        email_membership = email_membership.scalar_one_or_none()
        if email_membership:
            return email_membership 
        
        return False
    
    async def _active_admin_count(self, db: AsyncSession, organization: Organization, current_user: User) -> int:
        """Count active users holding full_admin_access in the org (via RBAC resolver)."""
        from app.core.permission_resolver import resolve_permissions, FULL_ADMIN

        memberships = await self.get_members(db, organization, current_user)
        count = 0
        for m in memberships:
            if not m.user_id:
                continue
            resolved = await resolve_permissions(db, str(m.user_id), str(organization.id))
            if FULL_ADMIN in resolved.org_permissions:
                count += 1
        return count
    
    async def _send_invitation_email(self, email: str, token: Optional[str] = None) -> str:
        """Send the invite email now, reliably. Returns "sent" or "failed".

        Awaited (not fire-and-forget) so the caller knows the real outcome,
        with a couple of retries for transient SMTP blips and a per-attempt
        timeout so a hung relay can't stall the invite request. The link carries
        the invite token (proof of inbox ownership at registration).
        """
        from urllib.parse import quote
        from app.services.notification_service import notification_service
        from app.services.email_copy import invite_email

        params = f"email={quote(email)}"
        if token:
            params = f"token={quote(token)}&{params}"
        sign_up_url = f"{settings.dash_config.base_url}/users/sign-up?{params}"
        subject, body = invite_email(sign_up_url)
        result = await notification_service.send_custom_email(
            recipients=[email],
            subject=subject,
            body=body,
            subtype="plain",
            retries=2,
            timeout=15,
        )
        if result.status != "sent":
            logger.error("Invitation email to %s failed: %s", email, result.error)
        return result.status


    async def get_organization_members(self, db: AsyncSession, current_user: User, organization: Organization) -> List[UserSchema]:
        # should get list of users via membership table
        result = await db.execute(select(Membership).where(Membership.organization_id == organization.id))
        memberships = result.scalars().all()
        user_ids = [membership.user_id for membership in memberships if membership.user_id is not None]
        result = await db.execute(select(User).where(User.id.in_(user_ids)))
        users = result.scalars().all()
        return [UserSchema.from_orm(user) for user in users]

    # ------------------------------------------------------------------
    # Excel / CSV import of memberships
    # ------------------------------------------------------------------

    MEMBERSHIP_IMPORT_MAX_ROWS = 1000
    _EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

    async def import_members(
        self,
        db: AsyncSession,
        organization: Organization,
        file_bytes: bytes,
        filename: str,
        dry_run: bool,
        current_user: User,
    ) -> MembershipImportReport:
        """Import memberships from .xlsx or .csv.

        Columns: `email` (required), `note` (optional). Other columns are ignored.
        Re-import is additive: existing roles/groups/status are never touched;
        only the `note` field is overwritten. Invites are not re-sent for rows
        whose email already has a (pending or active) membership.
        """
        rows = self._parse_membership_import(file_bytes, filename)
        report_rows: List[MembershipImportRow] = []
        summary = MembershipImportSummary()

        # Enterprise license seat cap. We track a running projection of how many *new*
        # members the import would create so the limit is reported up front — in the
        # dry-run preview and the real run alike — instead of failing row-by-row only
        # once writes start. -1 means unlimited. Existing members (updated/unchanged)
        # don't consume a new seat, so only freshly-created rows count toward it.
        from app.ee.license import get_max_users
        max_users = get_max_users()
        current_members = await self._count_org_memberships(db, organization.id) if max_users >= 0 else 0
        projected_new = 0

        for idx, raw in enumerate(rows, start=2):  # data starts at row 2 (header is row 1)
            email = (raw.get("email") or "").strip()
            note = raw.get("note")
            if note is not None:
                note = str(note).strip() or None
                if note and len(note) > MEMBERSHIP_NOTE_MAX_LENGTH:
                    report_rows.append(MembershipImportRow(
                        row=idx, email=email, note=note, status="error",
                        error=f"Note exceeds {MEMBERSHIP_NOTE_MAX_LENGTH} characters",
                    ))
                    summary.errors += 1
                    continue

            if not email:
                report_rows.append(MembershipImportRow(
                    row=idx, email=None, status="error", error="Missing email",
                ))
                summary.errors += 1
                continue
            if not self._EMAIL_RE.match(email):
                report_rows.append(MembershipImportRow(
                    row=idx, email=email, status="error", error="Invalid email format",
                ))
                summary.errors += 1
                continue

            existing = await self._find_membership_by_email(db, email, organization.id)
            if existing:
                if (existing.note or None) == (note or None):
                    report_rows.append(MembershipImportRow(
                        row=idx, email=email, note=note, status="unchanged",
                    ))
                    summary.unchanged += 1
                else:
                    if not dry_run:
                        existing.note = note
                        await db.flush()
                    report_rows.append(MembershipImportRow(
                        row=idx, email=email, note=note, status="updated",
                    ))
                    summary.updated += 1
                continue

            # New email — invite as a new pending membership.
            # Gate against the license seat cap using the running projection so the
            # overflow is reported identically in dry-run and real runs.
            if max_users >= 0 and current_members + projected_new >= max_users:
                report_rows.append(MembershipImportRow(
                    row=idx, email=email, note=note, status="error",
                    error=(
                        f"User limit reached for your license ({max_users}). "
                        "Contact sales to increase your seat count."
                    ),
                ))
                summary.errors += 1
                continue
            projected_new += 1

            if dry_run:
                report_rows.append(MembershipImportRow(
                    row=idx, email=email, note=note, status="created",
                ))
                summary.created += 1
            else:
                try:
                    await self.add_member(
                        db,
                        MembershipCreate(
                            organization_id=organization.id,
                            email=email,
                            role="member",
                            note=note,
                        ),
                        current_user,
                    )
                    report_rows.append(MembershipImportRow(
                        row=idx, email=email, note=note, status="created",
                    ))
                    summary.created += 1
                except HTTPException as e:
                    report_rows.append(MembershipImportRow(
                        row=idx, email=email, note=note, status="error",
                        error=str(e.detail),
                    ))
                    summary.errors += 1

        if not dry_run:
            await db.commit()

        return MembershipImportReport(dry_run=dry_run, summary=summary, rows=report_rows)

    async def _find_membership_by_email(self, db: AsyncSession, email: str, organization_id: str) -> Optional[Membership]:
        """Find a membership in this org by email, whether the user has registered or not."""
        user_q = await db.execute(select(User).where(User.email == email))
        user = user_q.scalar_one_or_none()
        if user:
            m_q = await db.execute(
                select(Membership).where(
                    Membership.user_id == user.id,
                    Membership.organization_id == organization_id,
                )
            )
            membership = m_q.scalar_one_or_none()
            if membership:
                return membership

        m_q = await db.execute(
            select(Membership).where(
                Membership.email == email,
                Membership.organization_id == organization_id,
            )
        )
        return m_q.scalar_one_or_none()

    def _parse_membership_import(self, file_bytes: bytes, filename: str) -> List[dict]:
        """Parse .xlsx or .csv into a list of {email, note} dicts.

        Header row is required. Column matching is case-insensitive and
        whitespace-trimmed. Columns other than ``email`` and ``note`` are
        ignored. Raises HTTPException(400) on unparseable input or empty file.
        """
        name = (filename or "").lower()
        ext = name.rsplit(".", 1)[-1] if "." in name else ""

        if ext in ("xlsx", "xlsm"):
            try:
                from openpyxl import load_workbook
            except ImportError as e:
                raise HTTPException(status_code=500, detail=f"Excel parsing unavailable: {e}")
            try:
                wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Failed to read xlsx: {e}")
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                raise HTTPException(status_code=400, detail="Empty file")
            header_map = self._build_header_map(header)
            self._require_email_column(header_map)
            data_rows = []
            for raw in rows_iter:
                if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in raw):
                    continue
                data_rows.append({key: raw[idx] if idx < len(raw) else None for key, idx in header_map.items()})
                if len(data_rows) >= self.MEMBERSHIP_IMPORT_MAX_ROWS:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Import exceeds max {self.MEMBERSHIP_IMPORT_MAX_ROWS} rows",
                    )
            return data_rows

        if ext == "csv" or not ext:
            try:
                text = file_bytes.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = file_bytes.decode("latin-1")
            reader = csv.reader(io.StringIO(text))
            try:
                header = next(reader)
            except StopIteration:
                raise HTTPException(status_code=400, detail="Empty file")
            header_map = self._build_header_map(header)
            self._require_email_column(header_map)
            data_rows = []
            for raw in reader:
                if not any((c or "").strip() for c in raw):
                    continue
                data_rows.append({key: raw[idx] if idx < len(raw) else None for key, idx in header_map.items()})
                if len(data_rows) >= self.MEMBERSHIP_IMPORT_MAX_ROWS:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Import exceeds max {self.MEMBERSHIP_IMPORT_MAX_ROWS} rows",
                    )
            return data_rows

        raise HTTPException(status_code=400, detail=f"Unsupported file type: .{ext}")

    @staticmethod
    def _build_header_map(header_row) -> dict:
        """Map known column names (case-insensitive) to their column indexes."""
        mapping = {}
        for idx, name in enumerate(header_row or []):
            if name is None:
                continue
            key = str(name).strip().lower()
            if key in ("email", "note") and key not in mapping:
                mapping[key] = idx
        return mapping

    @staticmethod
    def _require_email_column(header_map: dict) -> None:
        if "email" not in header_map:
            raise HTTPException(
                status_code=400,
                detail="Missing required 'email' column in header row",
            )
