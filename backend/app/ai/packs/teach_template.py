"""Teach Box templates — hand a user a fill-in-the-blank file, take it back.

Three public helpers (a route imports them):

  build_template_md()  -> str    : a labeled Markdown template with one filled
                                    CRM-lead example per block.
  build_template_xlsx() -> bytes : an .xlsx workbook (one tab per surface) with
                                    bold headers + one sample row each.
  parse_upload(name, raw) -> str : turn an uploaded .md/.txt/.csv/.xlsx back into
                                    plain text the classify prompt can chew on.

The block shapes mirror what `teach._CLASSIFY_PROMPT` expects (METRICS / QUERIES
/ EXAMPLES / DATA RULES / SKILLS / INSTRUCTIONS / KNOWLEDGE), so a filled-in
template round-trips straight through classify -> apply_spans.
"""

from __future__ import annotations

import io

_MAX_PARSE_CHARS = 20000


def build_template_md() -> str:
    """Markdown template — fill any block, delete the rest, upload back."""
    return """<!-- CityAgent teach template: fill any block, delete the rest, upload back. -->

## METRICS
- name: Total leads
  value: 1544
  logic: count rows where Related Brand Relationship Type = Lead
  sql: SELECT COUNT(*) FROM crm WHERE relationship_type = 'Lead'

## QUERIES
- name: Leads by channel
  sql: SELECT channel, COUNT(*) AS leads FROM crm WHERE relationship_type = 'Lead' GROUP BY channel ORDER BY leads DESC

## EXAMPLES
- question: How many leads did we get?
  answer: 1544 leads across all channels for the period.
  sql: SELECT COUNT(*) FROM crm WHERE relationship_type = 'Lead'

## DATA RULES
- lead = Related Brand Relationship Type = Lead
- fiscal year starts in April

## SKILLS
- name: Lead funnel summary
  when: user asks for a lead / conversion funnel overview
  steps: 1) count leads by channel 2) compute conversion vs successful 3) present a ranked table + short narrative

## INSTRUCTIONS
- Always report numbers with the period they cover.

## KNOWLEDGE
- The CRM covers retail membership and project tracking, six months of data (April-September).
"""


def build_template_xlsx() -> bytes:
    """.xlsx workbook, one tab per surface, bold header + one sample row."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except Exception as exc:  # pragma: no cover - env without openpyxl
        raise RuntimeError(
            "openpyxl is required to build the .xlsx teach template "
            "(pip install openpyxl)"
        ) from exc

    # (sheet name, [headers], [sample row])
    tabs = [
        ("Metrics",
         ["name", "value", "logic", "sql"],
         ["Total leads", 1544,
          "count rows where Related Brand Relationship Type = Lead",
          "SELECT COUNT(*) FROM crm WHERE relationship_type = 'Lead'"]),
        ("Queries",
         ["name", "sql"],
         ["Leads by channel",
          "SELECT channel, COUNT(*) AS leads FROM crm "
          "WHERE relationship_type = 'Lead' GROUP BY channel ORDER BY leads DESC"]),
        ("Examples",
         ["question", "answer", "sql"],
         ["How many leads did we get?",
          "1544 leads across all channels for the period.",
          "SELECT COUNT(*) FROM crm WHERE relationship_type = 'Lead'"]),
        ("DataRules",
         ["rule"],
         ["lead = Related Brand Relationship Type = Lead"]),
        ("Skills",
         ["name", "when", "steps"],
         ["Lead funnel summary",
          "user asks for a lead / conversion funnel overview",
          "1) count leads by channel 2) compute conversion 3) ranked table + narrative"]),
        ("Instructions",
         ["rule"],
         ["Always report numbers with the period they cover."]),
        ("Knowledge",
         ["fact"],
         ["The CRM covers retail membership and project tracking, "
          "six months of data (April-September)."]),
    ]

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    bold = Font(bold=True)
    for name, headers, sample in tabs:
        ws = wb.create_sheet(title=name)
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = bold
        ws.append(sample)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_upload(filename: str, raw: bytes) -> str:
    """Dispatch by extension -> plain text (capped at 20000 chars).

    .md/.txt/.markdown/.csv -> decoded text.
    .xlsx -> every sheet flattened to "## <Sheet>" + "key: val; key: val" lines.
    Unsupported extension -> ValueError.
    """
    name = (filename or "").lower()
    ext = name.rsplit(".", 1)[-1] if "." in name else ""

    if ext in ("md", "txt", "markdown"):
        text = raw.decode("utf-8", "ignore")
    elif ext == "csv":
        text = raw.decode("utf-8", "ignore")
    elif ext == "xlsx":
        text = _parse_xlsx(raw)
    else:
        raise ValueError(f"Unsupported teach-template file type: {filename!r}")

    return text[:_MAX_PARSE_CHARS]


def _parse_xlsx(raw: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - env without openpyxl
        raise RuntimeError(
            "openpyxl is required to parse an .xlsx teach template "
            "(pip install openpyxl)"
        ) from exc

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        keys = [str(h).strip() if h is not None else "" for h in header]
        out.append(f"## {ws.title}")
        for row in rows:
            cells = list(row) if row is not None else []
            pairs = []
            for i, key in enumerate(keys):
                val = cells[i] if i < len(cells) else None
                if val is None or str(val).strip() == "":
                    continue
                label = key or f"col{i + 1}"
                pairs.append(f"{label}: {str(val).strip()}")
            if pairs:
                out.append("; ".join(pairs))
    return "\n".join(out)
